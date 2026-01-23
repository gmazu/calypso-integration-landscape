from __future__ import annotations

import argparse
import ast
import sys
import textwrap
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
import uuid
import os

from manim import *
from openpyxl import load_workbook


# =============================================================================
# Funciones auxiliares
# =============================================================================
def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%y")
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%y")
        except ValueError:
            continue
    return text


def format_percent(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        pct = value * 100 if 0 <= value <= 1 else value
        return f"{int(round(pct))}%"
    text = str(value).strip()
    if not text:
        return ""
    return text if "%" in text else f"{text}%"


def load_tasks_from_xlsx(path: Path) -> list[list]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in header_row]
    header_map = {str(val).strip().lower(): idx for idx, val in enumerate(headers, start=1) if val}

    def col_index(*names: str) -> int | None:
        for name in names:
            key = name.strip().lower()
            if key in header_map:
                return header_map[key]
        return None

    col_id = col_index("id", "uid", "uid ")
    col_name = col_index("nombre de la tarea", "nombre", "task name", "name")
    col_status = col_index("estado", "status")
    col_assigned = col_index("asignado a", "asignado", "assigned")
    col_start = col_index("fecha de inicio", "inicio", "start", "start date")
    col_end = col_index("fecha de finalización", "fecha de finalizacion", "fin", "finish", "end", "end date")
    col_pct = col_index("porcentaje completo", "% completo", "avance", "percent complete", "percentcomplete")
    col_duration = col_index("duración", "duracion", "duration")
    col_pred = col_index("predecesores", "predecessors", "pred")

    if not col_name:
        raise ValueError("No se encontro la columna 'Nombre de la tarea' en el XLSX.")

    tasks = []
    for row_idx in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=col_name)
        name_val = name_cell.value
        if name_val is None or str(name_val).strip() == "":
            continue

        indent = name_cell.alignment.indent or 0
        level = int(indent)

        task_id = ws.cell(row=row_idx, column=col_id).value if col_id else row_idx - 1
        status = ws.cell(row=row_idx, column=col_status).value if col_status else ""
        assigned = ws.cell(row=row_idx, column=col_assigned).value if col_assigned else ""
        start = ws.cell(row=row_idx, column=col_start).value if col_start else ""
        end = ws.cell(row=row_idx, column=col_end).value if col_end else ""
        pct = ws.cell(row=row_idx, column=col_pct).value if col_pct else ""
        duration = ws.cell(row=row_idx, column=col_duration).value if col_duration else ""
        pred = ws.cell(row=row_idx, column=col_pred).value if col_pred else ""

        if isinstance(task_id, float) and task_id.is_integer():
            task_id = int(task_id)

        tasks.append(
            [
                task_id,
                level,
                str(name_val).strip(),
                str(status).strip() if status is not None else "",
                str(assigned).strip() if assigned is not None else "",
                format_date(start),
                format_date(end),
                format_percent(pct),
                str(duration).strip() if duration is not None else "",
                str(pred).strip() if pred is not None else "",
            ]
        )

    return tasks


def _parse_levels(values: list[str] | None) -> list[int]:
    if not values:
        return []
    levels: list[int] = []
    for raw in values:
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        for part in parts:
            levels.append(int(part))
    return levels


def filter_by_levels(tasks: list[list], levels: list[int]) -> list[list]:
    """Filtra por niveles e incluye ancestros para dar contexto visual."""
    if not levels:
        return tasks
    level_set = set(levels)
    result: list[list] = []
    seen: set[tuple] = set()
    stack: list[list] = []

    for row in tasks:
        level = row[1]
        while stack and stack[-1][1] >= level:
            stack.pop()
        stack.append(row)

        if level in level_set:
            # Incluir ancestros y la fila actual sin duplicados.
            for parent in stack:
                key = tuple(parent)
                if key not in seen:
                    result.append(parent)
                    seen.add(key)

    return result


def parse_filter_sequence(argv: list[str]) -> list[tuple[str, list[int] | int]]:
    """Parsea --nivel/--id en el orden recibido para aplicar filtros anidados."""
    seq: list[tuple[str, list[int] | int]] = []
    i = 0
    while i < len(argv):
        arg = argv[i]
        if arg == "--nivel":
            if i + 1 >= len(argv):
                raise SystemExit("Error: --nivel requiere un valor.")
            levels = _parse_levels([argv[i + 1]])
            seq.append(("nivel", levels))
            i += 2
            continue
        if arg == "--id":
            if i + 1 >= len(argv):
                raise SystemExit("Error: --id requiere un valor.")
            try:
                task_id = int(argv[i + 1])
            except ValueError as exc:
                raise SystemExit("Error: --id requiere un entero.") from exc
            seq.append(("id", task_id))
            i += 2
            continue
        i += 1
    return seq


def split_by_pipe(argv: list[str]) -> list[list[str]]:
    """Divide argumentos en segmentos separados por '|'."""
    segments: list[list[str]] = []
    current: list[str] = []
    for arg in argv:
        if arg == "|":
            segments.append(current)
            current = []
        else:
            current.append(arg)
    segments.append(current)
    return segments


def filter_by_id_with_context(tasks: list[list], task_id: int, max_depth: int | None = None) -> list[list]:
    """
    Filtra la tarea con el ID dado más sus hijos directos.
    Retorna la tarea padre y todas las tareas con nivel mayor hasta encontrar
    otra tarea del mismo nivel o menor.
    """
    result: list[list] = []
    seen: set[tuple] = set()
    stack: list[list] = []

    found_idx = None
    parent_level = None

    for idx, row in enumerate(tasks):
        level = row[1]
        while stack and stack[-1][1] >= level:
            stack.pop()
        stack.append(row)
        if row[0] == task_id:
            found_idx = idx
            parent_level = level
            for parent in stack:
                key = tuple(parent)
                if key not in seen:
                    result.append(parent)
                    seen.add(key)
            break

    if found_idx is None or parent_level is None:
        return []

    # Agregar hijos (nivel > parent_level) hasta encontrar mismo nivel o menor
    for row in tasks[found_idx + 1:]:
        if row[1] <= parent_level:
            break
        if max_depth is not None and row[1] > parent_level + max_depth:
            continue
        key = tuple(row)
        if key not in seen:
            result.append(row)
            seen.add(key)

    return result


def get_tasks_for_render() -> list[list]:
    """Lee tareas desde filter_gantt.tasks (generado por el CLI)."""
    tasks_file = Path(__file__).with_name("filter_gantt.tasks")
    if not tasks_file.exists():
        raise FileNotFoundError(
            f"No se encontró {tasks_file}. "
            "Ejecuta primero el CLI para generar filter_gantt.tasks."
        )
    return load_tasks_from_file(tasks_file)


def load_tasks_from_file(path: Path) -> list[list]:
    text = path.read_text(encoding="utf-8").lstrip()
    module = ast.parse(text)
    for node in module.body:
        if isinstance(node, ast.Assign):
            for target in node.targets:
                if isinstance(target, ast.Name) and target.id == "tasks":
                    return ast.literal_eval(node.value)
    raise ValueError("tasks list not found in file")


def write_tasks_file(tasks: list[list], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8") as f:
        f.write("tasks = [\n")
        for row in tasks:
            f.write(f"    {row},\n")
        f.write("]\n")


# =============================================================================
# CLI standalone (python gantt_timeline_v2.py -xlsx ... --nivel ...)
# =============================================================================
def run_filter_cli() -> int:
    """CLI para generar filter_gantt.tasks desde XLSX."""
    parser = argparse.ArgumentParser(
        description="Genera filter_gantt.tasks desde XLSX o renderiza con Manim."
    )
    parser.add_argument("-xlsx", "--xlsx", required=True, type=Path, help="Ruta al archivo XLSX.")
    parser.add_argument(
        "--expand",
        action="store_true",
        help="Al usar --id, expande solo el siguiente nivel del ID desde el XLSX completo.",
    )
    parser.add_argument(
        "-debug",
        "--debug",
        action="store_true",
        help="Imprime un informe breve del filtro (IDs/niveles/títulos).",
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=Path(__file__).with_name("filter_gantt.tasks"),
        help="Archivo de salida (default: filter_gantt.tasks).",
    )
    args, _unknown = parser.parse_known_args()
    segments = split_by_pipe(sys.argv[1:])

    if not args.xlsx.exists():
        print(f"Error: no existe el archivo {args.xlsx}", file=sys.stderr)
        return 1

    filtered = load_tasks_from_xlsx(args.xlsx)
    temp_files: list[Path] = []

    for seg_idx, seg in enumerate(segments):
        filter_seq = parse_filter_sequence(seg)
        if filter_seq:
            for kind, value in filter_seq:
                if kind == "nivel":
                    filtered = filter_by_levels(
                        filtered, value if isinstance(value, list) else [value]
                    )
                    print(f"Filtrado por nivel {value}: {len(filtered)} tareas")
                elif kind == "id":
                    base = load_tasks_from_xlsx(args.xlsx) if args.expand else filtered
                    depth = 1 if args.expand else None
                    filtered = filter_by_id_with_context(base, int(value), max_depth=depth)
                    print(f"Filtrado por ID {value}: {len(filtered)} tareas")
        else:
            if seg_idx == 0:
                print(f"Sin filtro: {len(filtered)} tareas")

        if seg_idx < len(segments) - 1:
            tmp = Path("/tmp") / f"gantt_filter_{uuid.uuid4().hex}_{seg_idx}.tasks"
            write_tasks_file(filtered, tmp)
            temp_files.append(tmp)
            filtered = load_tasks_from_file(tmp)

    if args.debug:
        levels = sorted({row[1] for row in filtered})
        print(f"Niveles presentes: {levels}")
        print("Tareas filtradas (id | nivel | título):")
        for task_id, level, name, *_rest in filtered:
            print(f"- {task_id} | {level} | {name}")

    write_tasks_file(filtered, args.output)
    print(f"Escrito: {args.output}")

    for tmp in temp_files:
        try:
            os.remove(tmp)
        except OSError:
            pass
    return 0


# =============================================================================
# Escenas Manim
# =============================================================================

# Uso:
#   manim -pql gantt_timeline_v2.py GanttTimelineLevel2 --xlsx archivo.xlsx --nivel 1
#   manim -pql gantt_timeline_v2.py GanttTimelineLevel2 --xlsx archivo.xlsx --id 42


class GanttTimelineLevel2(Scene):
    def construct(self):
        tasks = get_tasks_for_render()

        title_text = "Hablitación Plataforma Calypso Banco BCI"
        subtitle_text = "Ambiente Pre Productivo"
        level0 = next((row for row in tasks if row[1] == 0), None)
        level1 = next((row for row in tasks if row[1] == 1), None)
        if level0:
            title_text = level0[2]
        if level1:
            subtitle_text = level1[2]

        tasks = [row for row in tasks if row[1] >= 2]

        dated = []
        undated = []
        for row in tasks:
            task_id, _, name, *_rest, start, end, pct, _dur, _pred = row
            if start and end:
                dated.append(
                    {
                        "id": task_id,
                        "name": name,
                        "start": datetime.strptime(start, "%d/%m/%y"),
                        "end": datetime.strptime(end, "%d/%m/%y"),
                        "start_str": start,
                        "end_str": end,
                        "pct": pct,
                    }
                )
            else:
                undated.append({"id": task_id, "name": name})

        dated.sort(key=lambda t: t["start"])

        title = Text(title_text, font_size=28, weight=BOLD)
        subtitle = Text(subtitle_text, font_size=16, color=GRAY_B)
        header = VGroup(title, subtitle).arrange(DOWN, buff=0.2).to_edge(UP, buff=0.4)

        timeline_left = LEFT * 5.5 + DOWN * 0.2
        timeline_right = RIGHT * 5.5 + DOWN * 0.2
        timeline = Line(timeline_left, timeline_right, color=GRAY_B, stroke_width=4)

        if dated:
            start_min = min(t["start"] for t in dated)
            end_max = max(t["start"] for t in dated)
        else:
            start_min = datetime.now()
            end_max = datetime.now()

        def date_to_x(value: datetime) -> float:
            total = (end_max - start_min).days or 1
            offset = (value - start_min).days
            ratio = offset / total
            return interpolate(timeline_left[0], timeline_right[0], ratio)

        scale_y = timeline_left[1] - 3.2

        # Fecha de hoy (ajuste de año solo si cae dentro del rango del Gantt)
        today = datetime.now()
        if today.year != start_min.year:
            try:
                candidate = today.replace(year=start_min.year)
            except ValueError:
                candidate = today.replace(year=start_min.year, day=28)
            if start_min <= candidate <= end_max:
                today = candidate

        # Promedio global de avance real y planificado para marcador de "hoy"
        pct_all = []
        for t in dated:
            if t.get("pct"):
                try:
                    pct_all.append(float(str(t["pct"]).replace("%", "").strip()))
                except ValueError:
                    pass
        avg_all = round(sum(pct_all) / len(pct_all)) if pct_all else None
        planned_all = []
        for t in dated:
            start_d = t["start"]
            end_d = t["end"]
            total_days = max(1, (end_d - start_d).days)
            if today <= start_d:
                planned = 0.0
            elif today >= end_d:
                planned = 100.0
            else:
                elapsed = (today - start_d).days
                planned = (elapsed / total_days) * 100.0
            planned_all.append(planned)
        avg_planned = round(sum(planned_all) / len(planned_all)) if planned_all else None

        # Línea de "hoy" interpolada entre puntos vecinos para respetar separaciones reales
        date_keys = [t["start"].date() for t in dated]
        if date_keys:
            today_date = today.date()
            prev_d = max((d for d in date_keys if d <= today_date), default=date_keys[0])
            next_d = min((d for d in date_keys if d >= today_date), default=date_keys[-1])
            x_prev = date_to_x(datetime.combine(prev_d, datetime.min.time()))
            x_next = date_to_x(datetime.combine(next_d, datetime.min.time()))
            span = (next_d - prev_d).days or 1
            ratio_local = (today_date - prev_d).days / span
            x_today = x_prev + (x_next - x_prev) * ratio_local
        else:
            x_today = date_to_x(today)
        # Dial vintage: barras paralelas para real vs plan (más separación si difieren)
        dial_height = 0.55
        dial_w = 0.08
        plan_val = avg_planned if avg_planned is not None else 0
        real_val = avg_all if avg_all is not None else 0
        diff = abs((real_val or 0) - (plan_val or 0))
        dial_gap = 0.02 + 0.0015 * diff
        dial_gap = min(0.22, dial_gap)
        dial_real = Rectangle(
            width=dial_w,
            height=dial_height,
            stroke_width=0,
            fill_color=RED_E,
            fill_opacity=0.35,
        ).move_to([x_today - dial_gap / 2, scale_y + dial_height / 2, 0])
        dial_plan = Rectangle(
            width=dial_w,
            height=dial_height,
            stroke_width=0,
            fill_color=RED_A,
            fill_opacity=0.35,
        ).move_to([x_today + dial_gap / 2, scale_y + dial_height / 2, 0])
        today_line = VGroup(dial_real, dial_plan)
        today_tick = Line(
            [x_today, timeline_left[1] - 0.18, 0],
            [x_today, timeline_left[1] + 0.18, 0],
            color=RED,
            stroke_width=1,
        )
        today_label = Text(f"Hoy {today.strftime('%d/%m')}", font_size=10, color=RED)
        today_label.next_to(today_line, UP, buff=0.04)
        pct_parts = []
        if avg_all is not None:
            pct_parts.append(f"Real {avg_all}%")
        if avg_planned is not None:
            pct_parts.append(f"Plan {avg_planned}%")
        if pct_parts:
            today_pct = Text(" | ".join(pct_parts), font_size=10, color=RED)
            today_pct.next_to(today_label, RIGHT, buff=0.15)
        else:
            today_pct = None
        if "DEBUG_TODAY" in os.environ:
            print(f"[DEBUG_TODAY] start_min={start_min.date()} end_max={end_max.date()} today={today.date()}")

        points = VGroup()
        stems_bg = VGroup()
        stems_lit = VGroup()
        labels = VGroup()
        dates = VGroup()
        deltas = VGroup()
        full_test = VGroup()
        pct_by_date: dict = {}

        grouped = OrderedDict()
        for task in dated:
            key = task["start"].date()
            grouped.setdefault(key, []).append(task)

        above_idx = 0
        below_idx = 0
        spacing_scale = 0.85

        date_keys = list(grouped.keys())
        for idx, (key, tasks_for_date) in enumerate(grouped.items()):
            x = date_to_x(tasks_for_date[0]["start"])
            y = timeline_left[1]

            point = Dot([x, y, 0], radius=0.065, color=BLUE_D)

            above = idx % 2 == 0
            if above:
                stem_len = [1.0, 1.5, 2.0][above_idx % 3] * spacing_scale
                above_idx += 1
            else:
                stem_len = [1.0, 1.5, 2.0][below_idx % 3] * spacing_scale
                below_idx += 1

            date_text = tasks_for_date[0]["start"].strftime("%d/%m")
            date_label = Text(date_text, font_size=12, color=BLUE_D)
            date_label.next_to(point, DOWN if above else UP, buff=0.1)

            offsets = [1.0, 1.5, 2.0]
            if len(tasks_for_date) > 1:
                offsets = [o * spacing_scale for o in offsets]
            else:
                offsets = [stem_len + 0.2 * spacing_scale]
            x_offsets = [0.45, -0.45, 0.9, -0.9, 1.35, -1.35]
            for t_idx, task in enumerate(tasks_for_date):
                offset = offsets[t_idx] if t_idx < len(offsets) else offsets[-1]
                target_y = y + (offset if above else -offset)

                id_line = f"ID {task['id']}"
                if task.get("pct"):
                    id_line = f"{id_line}  {task['pct']}"
                title_text = Text(id_line, font_size=12, weight=BOLD)
                end_text = Text(f"Fin: {task['end_str']}", font_size=11, color=GRAY_C)
                text_block = VGroup(title_text, end_text).arrange(DOWN, buff=0.06, aligned_edge=LEFT)

                text_block.move_to([x, target_y, 0])
                x_shift = x_offsets[t_idx] if t_idx < len(x_offsets) else x_offsets[-1]
                text_block.shift(RIGHT * x_shift)
                if target_y >= y:
                    text_block.shift(UP * 0.35)
                else:
                    text_block.shift(DOWN * 0.35)
                labels.add(text_block)

                # Barra vertical tipo ecualizador (segmentos horizontales)
                bar_width = 0.18
                bar_height = abs(target_y - y) * 0.9
                bar_center = (y + target_y) / 2

                pct_val = None
                if task.get("pct"):
                    try:
                        pct_val = float(str(task["pct"]).replace("%", "").strip())
                    except ValueError:
                        pct_val = None
                # Siempre dibujar segmentos de fondo (apagados)
                seg_count = 14
                gap_ratio = 0.2
                seg_height = bar_height / (seg_count + (seg_count - 1) * gap_ratio)
                seg_gap = seg_height * gap_ratio
                fill_bg = VGroup()
                fill_lit = VGroup()
                halo = Rectangle(
                    width=bar_width,
                    height=bar_height,
                    stroke_width=0,
                    fill_color=GRAY_E,
                    fill_opacity=0.12,
                ).move_to([x, bar_center, 0])
                fill_bg.add(halo)
                for s in range(seg_count):
                    seg = Rectangle(
                        width=bar_width,
                        height=seg_height,
                        stroke_width=0,
                        fill_color=GRAY_C,
                        fill_opacity=0.28,
                    )
                    if target_y >= y:
                        seg_y = y + (seg_height / 2) + s * (seg_height + seg_gap)
                    else:
                        seg_y = y - (seg_height / 2) - s * (seg_height + seg_gap)
                    seg.move_to([x, seg_y, 0])
                    fill_bg.add(seg)

                # Encender segmentos según % (si hay valor)
                if pct_val is not None and pct_val > 0:
                    pct_norm = max(0.0, min(1.0, pct_val / 100.0))
                    lit_segments = int(round(pct_norm * seg_count))
                    if pct_norm > 0 and lit_segments == 0:
                        lit_segments = 1
                    for s in range(lit_segments):
                        t = s / max(1, seg_count - 1)
                        if t <= 0.5:
                            color = interpolate_color(RED_E, GREEN_B, t * 2)
                        else:
                            color = interpolate_color(GREEN_B, BLUE_E, (t - 0.5) * 2)
                        seg = Rectangle(
                            width=bar_width,
                            height=seg_height,
                            stroke_width=0,
                            fill_color=color,
                            fill_opacity=0.95,
                        )
                        if target_y >= y:
                            seg_y = y + (seg_height / 2) + s * (seg_height + seg_gap)
                        else:
                            seg_y = y - (seg_height / 2) - s * (seg_height + seg_gap)
                        seg.move_to([x, seg_y, 0])
                        fill_lit.add(seg)

                # Overlay para prueba de calidad: 100% lleno (se anima al final)
                full_fill = VGroup()
                for s in range(seg_count):
                    t = s / max(1, seg_count - 1)
                    if t <= 0.5:
                        color = interpolate_color(RED_E, GREEN_B, t * 2)
                    else:
                        color = interpolate_color(GREEN_B, BLUE_E, (t - 0.5) * 2)
                    seg = Rectangle(
                        width=bar_width,
                        height=seg_height,
                        stroke_width=0,
                        fill_color=color,
                        fill_opacity=0.8,
                    )
                    if target_y >= y:
                        seg_y = y + (seg_height / 2) + s * (seg_height + seg_gap)
                    else:
                        seg_y = y - (seg_height / 2) - s * (seg_height + seg_gap)
                    seg.move_to([x, seg_y, 0])
                    full_fill.add(seg)
                full_test.add(full_fill)
                # Cap para 100%: asegura que llegue al borde
                if pct_val is not None and pct_val >= 99.9:
                    cap_h = seg_height * 0.6
                    cap = Rectangle(
                        width=bar_width,
                        height=cap_h,
                        stroke_width=0,
                        fill_color=YELLOW_B,
                        fill_opacity=0.95,
                    )
                    if target_y >= y:
                        cap.move_to([x, y + bar_height - cap_h / 2, 0])
                    else:
                        cap.move_to([x, y - bar_height + cap_h / 2, 0])
                    fill_lit.add(cap)

                stems_bg.add(fill_bg)
                if len(fill_lit) > 0:
                    stems_lit.add(fill_lit)

            # Marcas de escala (0-100) junto a la barra (solo una vez por fecha)
            scale_marks = VGroup()
            ticks = [0, 25, 50, 75, 100]
            for t in ticks:
                frac = t / 100.0
                tick_len = 0.16 if t in (0, 50, 100) else 0.08
                if above:
                    ty = y + frac * stem_len
                else:
                    ty = y - frac * stem_len
                scale_marks.add(
                    Line([x - 0.18, ty, 0], [x - 0.18 - tick_len, ty, 0], color=GRAY_C, stroke_width=1)
                )
                if t in (0, 50, 100):
                    lbl = Text(str(t), font_size=9, color=GRAY_C)
                    lbl.next_to(scale_marks[-1], LEFT, buff=0.04)
                    scale_marks.add(lbl)

            points.add(point)
            dates.add(date_label)
            stems_bg.add(scale_marks)

            pcts = []
            for t in tasks_for_date:
                if t.get("pct"):
                    try:
                        pcts.append(float(str(t["pct"]).replace("%", "").strip()))
                    except ValueError:
                        pass
            if pcts:
                pct_by_date[key] = round(sum(pcts) / len(pcts))

        # Escala inferior estilo "mapa": barra segmentada con dias por tramo
        bar_height = 0.15
        bar_bg = VGroup()
        bar_lit = VGroup()
        bar_full = VGroup()
        for i in range(1, len(date_keys)):
            d0 = date_keys[i - 1]
            d1 = date_keys[i]
            delta_days = (d1 - d0).days
            x0 = date_to_x(datetime.combine(d0, datetime.min.time()))
            x1 = date_to_x(datetime.combine(d1, datetime.min.time()))
            mid_x = (x0 + x1) / 2

            seg_width = max(0.01, x1 - x0)
            days = max(1, delta_days)
            unit_w = seg_width / days
            pct_for_span = pct_by_date.get(d1)
            pct_norm = 1.0 if pct_for_span is None else max(0.0, min(1.0, pct_for_span / 100.0))
            for d in range(days):
                t_prog = d / max(1, days - 1)
                # Fondo apagado
                bg_seg = Rectangle(
                    width=unit_w * 0.85,
                    height=bar_height,
                    stroke_width=0,
                    fill_color=GRAY_C,
                    fill_opacity=0.32,
                )
                bg_seg.move_to([x0 + (d + 0.5) * unit_w, scale_y, 0])
                bar_bg.add(bg_seg)

                # Full test (100%)
                if t_prog <= 0.5:
                    full_color = interpolate_color(RED_E, GREEN_B, t_prog * 2)
                else:
                    full_color = interpolate_color(GREEN_B, BLUE_E, (t_prog - 0.5) * 2)
                full_seg = Rectangle(
                    width=unit_w * 0.85,
                    height=bar_height,
                    stroke_width=0,
                    fill_color=full_color,
                    fill_opacity=1,
                )
                full_seg.move_to([x0 + (d + 0.5) * unit_w, scale_y, 0])
                bar_full.add(full_seg)

                if t_prog <= pct_norm:
                    if t_prog <= 0.5:
                        color = interpolate_color(RED_E, GREEN_B, t_prog * 2)
                    else:
                        color = interpolate_color(GREEN_B, BLUE_E, (t_prog - 0.5) * 2)
                    opacity = 1
                else:
                    color = None
                    opacity = 0
                seg = Rectangle(
                    width=unit_w * 0.85,
                    height=bar_height,
                    stroke_width=0,
                    fill_color=color if color else GRAY_C,
                    fill_opacity=opacity,
                )
                seg_x = x0 + (d + 0.5) * unit_w
                seg.move_to([seg_x, scale_y, 0])
                if opacity > 0:
                    bar_lit.add(seg)

            tick = Line([x0, scale_y + 0.08, 0], [x0, scale_y - 0.08, 0], color=GRAY_B, stroke_width=1)
            txt = Text(f"{delta_days}d", font_size=9, color=GRAY_B)
            txt.move_to([mid_x, scale_y - 0.28, 0])
            if d1 in pct_by_date:
                avg_txt = Text(f"{pct_by_date[d1]}%", font_size=10, color=GREEN_C)
                avg_txt.move_to([mid_x, scale_y + 0.22, 0])
                deltas.add(VGroup(tick, avg_txt, txt))
            else:
                deltas.add(VGroup(tick, txt))

            # Fecha en el punto de separación del tramo
            date_str = d1.strftime("%d/%m")
            fs = max(7, min(9, seg_width * 3))
            date_lbl = Text(date_str, font_size=fs, color=GRAY_C)
            date_lbl.next_to(tick, DOWN, buff=0.03)
            deltas.add(date_lbl)

        if date_keys:
            x_start = date_to_x(datetime.combine(date_keys[0], datetime.min.time()))
            x_end = date_to_x(datetime.combine(date_keys[-1], datetime.min.time()))
            deltas.add(Line([x_end, scale_y + 0.08, 0], [x_end, scale_y - 0.08, 0], color=GRAY_B, stroke_width=1))

        if undated:
            undated_title = Text("Sin fechas", font_size=16, color=GRAY_B)
            undated_lines = VGroup(
                *[Text(f"{t['id']} - {t['name']}", font_size=14) for t in undated]
            ).arrange(DOWN, aligned_edge=LEFT, buff=0.2)
            undated_block = VGroup(undated_title, undated_lines).arrange(DOWN, buff=0.2)
            undated_block.to_edge(RIGHT, buff=0.6).shift(DOWN * 2.2)
        else:
            undated_block = VGroup()

        self.play(Write(header), run_time=1)
        self.play(Create(timeline), run_time=0.8)
        today_group = None
        if today_line:
            if today_pct:
                today_group = VGroup(today_line, today_tick, today_label, today_pct)
            else:
                today_group = VGroup(today_line, today_tick, today_label)
        self.play(LaggedStartMap(FadeIn, points, lag_ratio=0.05), run_time=0.9)
        self.play(LaggedStartMap(FadeIn, stems_bg, lag_ratio=0.05), run_time=1.0)
        self.play(LaggedStartMap(FadeIn, dates, lag_ratio=0.05), run_time=0.8)
        self.play(LaggedStartMap(FadeIn, labels, lag_ratio=0.05), run_time=1.2)
        if deltas:
            self.play(FadeIn(bar_bg), run_time=0.4)
            self.play(LaggedStartMap(FadeIn, deltas, lag_ratio=0.03), run_time=0.6)

        # Mostrar "hoy" junto con el resto de elementos
        if today_group:
            self.play(FadeIn(today_group), run_time=0.4)

        # Prueba de calidad: llenar ecualizadores y barra inferior al 100% brevemente
        if full_test:
            self.play(AnimationGroup(*[FadeIn(g) for g in full_test], lag_ratio=1), run_time=0.9)
            if bar_full:
                self.play(AnimationGroup(*[FadeIn(g) for g in bar_full], lag_ratio=0.05), run_time=0.7)
            self.wait(0.3)
            self.play(FadeOut(full_test), FadeOut(bar_full), run_time=0.4)

        # Mostrar valores reales después de la prueba
        if stems_lit:
            self.play(LaggedStartMap(FadeIn, stems_lit, lag_ratio=0.05), run_time=0.8)
        if bar_lit:
            self.play(FadeIn(bar_lit), run_time=0.4)

        if undated_block:
            self.play(FadeIn(undated_block), run_time=0.6)

        self.wait(2)


class GanttTimelineCircular(ThreeDScene):
    def construct(self):
        tasks = get_tasks_for_render()

        title_text = "Hablitación Plataforma Calypso Banco BCI"
        subtitle_text = "Ambiente Pre Productivo"
        level0 = next((row for row in tasks if row[1] == 0), None)
        level1 = next((row for row in tasks if row[1] == 1), None)
        if level0:
            title_text = level0[2]
        if level1:
            subtitle_text = level1[2]

        tasks = [row for row in tasks if row[1] >= 2]

        dated = []
        for row in tasks:
            task_id, _, name, *_rest, start, end, _pct, _dur, _pred = row
            if start and end:
                dated.append(
                    {
                        "id": task_id,
                        "name": name,
                        "start": datetime.strptime(start, "%d/%m/%y"),
                    }
                )

        dated.sort(key=lambda t: t["start"])

        title = Text(title_text, font_size=26, weight=BOLD)
        subtitle = Text(subtitle_text, font_size=16, color=GRAY_B)
        header = VGroup(title, subtitle).arrange(DOWN, buff=0.2).to_edge(UP, buff=0.4)

        outer = Arc(
            radius=3.2,
            start_angle=PI * 0.2,
            angle=PI * 1.6,
            color=GRAY_B,
            stroke_width=2,
        )
        inner = Arc(
            radius=2.6,
            start_angle=PI * 0.2,
            angle=PI * 1.6,
            color=GRAY_C,
            stroke_width=1,
        )

        ticks = VGroup()
        for i in range(70):
            ang = outer.start_angle + (outer.angle * i / 69)
            p1 = 3.25 * np.array([np.cos(ang), np.sin(ang), 0])
            p2 = 3.4 * np.array([np.cos(ang), np.sin(ang), 0])
            ticks.add(Line(p1, p2, stroke_width=1, color=GRAY_B))

        points = VGroup()
        labels = VGroup()
        if dated:
            start_min = min(t["start"] for t in dated)
            end_max = max(t["start"] for t in dated)
        else:
            start_min = datetime.now()
            end_max = datetime.now()

        def date_to_angle(value: datetime) -> float:
            total = (end_max - start_min).days or 1
            offset = (value - start_min).days
            ratio = offset / total
            return outer.start_angle + outer.angle * ratio

        for idx, task in enumerate(dated):
            ang = date_to_angle(task["start"])
            radius = 2.9
            pos = radius * np.array([np.cos(ang), np.sin(ang), 0])
            points.add(Dot(pos, radius=0.05, color=BLUE_D))

            label = Text(f"ID {task['id']}", font_size=12, color=GRAY_B)
            label.move_to((radius + 0.45) * np.array([np.cos(ang), np.sin(ang), 0]))
            label.rotate(ang + PI / 2)
            labels.add(label)

        group = VGroup(outer, inner, ticks, points, labels)
        group.move_to(DOWN * 0.2)
        group.rotate(PI / 5, axis=RIGHT)
        group.rotate(-PI / 10, axis=UP)

        self.set_camera_orientation(phi=70 * DEGREES, theta=-35 * DEGREES)

        self.play(Write(header), run_time=1)
        self.play(Create(outer), Create(inner), run_time=1.2)
        self.play(LaggedStartMap(Create, ticks, lag_ratio=0.02), run_time=1.5)
        self.play(LaggedStartMap(FadeIn, points, lag_ratio=0.05), run_time=1.0)
        self.play(LaggedStartMap(FadeIn, labels, lag_ratio=0.05), run_time=1.0)
        self.play(Rotate(group, angle=PI / 2, axis=UP), run_time=4, rate_func=linear)
        self.wait(1)


if __name__ == "__main__":
    raise SystemExit(run_filter_cli())
