from __future__ import annotations

from datetime import datetime
import ast
from pathlib import Path
from collections import OrderedDict
import textwrap
import argparse
import sys
import os

from manim import *
from openpyxl import load_workbook


DEFAULT_TASK_FILE = "Gantt/smartsheet_PRE_v1.txt"


def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%y")
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%y")
        except ValueError:
            continue
    return text


def format_percent(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        pct = value * 100 if 0 <= value <= 1 else value
        return f"{int(round(pct))}%"
    text = str(value).strip()
    if not text:
        return ""
    return text if "%" in text else f"{text}%"


def load_tasks_from_txt(path: Path) -> list[list]:
    text = path.read_text(encoding="utf-8")
    module = ast.parse(text)
    for node in module.body:
        if isinstance(node, ast.Assign):
            for target in node.targets:
                if isinstance(target, ast.Name) and target.id == "tasks":
                    return ast.literal_eval(node.value)
    raise ValueError("tasks list not found")


def load_tasks_from_xlsx(path: Path) -> list[list]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in header_row]
    header_map = {str(val).strip().lower(): idx for idx, val in enumerate(headers, start=1) if val}

    def col_index(*names: str) -> int | None:
        for name in names:
            key = name.strip().lower()
            if key in header_map:
                return header_map[key]
        return None

    col_id = col_index("id", "uid", "uid ")
    col_name = col_index("nombre de la tarea", "nombre", "task name", "name")
    col_status = col_index("estado", "status")
    col_assigned = col_index("asignado a", "asignado", "assigned")
    col_start = col_index("fecha de inicio", "inicio", "start", "start date")
    col_end = col_index("fecha de finalización", "fecha de finalizacion", "fin", "finish", "end", "end date")
    col_pct = col_index("porcentaje completo", "% completo", "avance", "percent complete", "percentcomplete")
    col_duration = col_index("duración", "duracion", "duration")
    col_pred = col_index("predecesores", "predecessors", "pred")

    if not col_name:
        raise ValueError("No se encontro la columna 'Nombre de la tarea' en el XLSX.")

    tasks = []
    for row_idx in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=col_name)
        name_val = name_cell.value
        if name_val is None or str(name_val).strip() == "":
            continue

        indent = name_cell.alignment.indent or 0
        level = int(indent)

        task_id = ws.cell(row=row_idx, column=col_id).value if col_id else row_idx - 1
        status = ws.cell(row=row_idx, column=col_status).value if col_status else ""
        assigned = ws.cell(row=row_idx, column=col_assigned).value if col_assigned else ""
        start = ws.cell(row=row_idx, column=col_start).value if col_start else ""
        end = ws.cell(row=row_idx, column=col_end).value if col_end else ""
        pct = ws.cell(row=row_idx, column=col_pct).value if col_pct else ""
        duration = ws.cell(row=row_idx, column=col_duration).value if col_duration else ""
        pred = ws.cell(row=row_idx, column=col_pred).value if col_pred else ""

        tasks.append(
            [
                task_id,
                level,
                str(name_val).strip(),
                str(status).strip() if status is not None else "",
                str(assigned).strip() if assigned is not None else "",
                format_date(start),
                format_date(end),
                format_percent(pct),
                str(duration).strip() if duration is not None else "",
                str(pred).strip() if pred is not None else "",
            ]
        )

    return tasks


def load_tasks(path: str) -> list[list]:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"No existe el archivo: {p}")

    if p.suffix.lower() in {".xlsx", ".xls"}:
        return load_tasks_from_xlsx(p)
    return load_tasks_from_txt(p)


# We need to parse our arguments before Manim does.
parser = argparse.ArgumentParser(description="Gantt chart timeline generator for Manim.")
parser.add_argument(
    "-xlsx",
    "--xlsx",
    dest="source",
    help="Path to the tasks data file (XLSX/TXT). If not specified, uses default.",
)
custom_args, _ = parser.parse_known_args()


class GanttTimelineLevel2(Scene):
    def construct(self):
        source_path = custom_args.source or os.environ.get("GANTT_SOURCE") or DEFAULT_TASK_FILE
        tasks = load_tasks(source_path)
        level2 = [row for row in tasks if row[1] == 2]

        dated = []
        undated = []
        for row in level2:
            task_id, _, name, *_rest, start, end, _pct, _dur, _pred = row
            if start and end:
                dated.append(
                    {
                        "id": task_id,
                        "name": name,
                        "start": datetime.strptime(start, "%d/%m/%y"),
                        "end": datetime.strptime(end, "%d/%m/%y"),
                        "start_str": start,
                        "end_str": end,
                    }
                )
            else:
                undated.append({"id": task_id, "name": name})

        dated.sort(key=lambda t: t["start"])

        title = Text("Hablitación Plataforma Calypso Banco BCI", font_size=34, weight=BOLD)
        subtitle = Text("Ambiente Pre Productivo", font_size=18, color=GRAY_B)
        header = VGroup(title, subtitle).arrange(DOWN, buff=0.2).to_edge(UP, buff=0.4)

        timeline_left = LEFT * 5.5 + DOWN * 0.2
        timeline_right = RIGHT * 5.5 + DOWN * 0.2
        timeline = Line(timeline_left, timeline_right, color=GRAY_B, stroke_width=4)

        if dated:
            start_min = min(t["start"] for t in dated)
            end_max = max(t["end"] for t in dated)
        else:
            start_min = datetime.now()
            end_max = datetime.now()

        def date_to_x(value: datetime) -> float:
            total = (end_max - start_min).days or 1
            offset = (value - start_min).days
            ratio = offset / total
            return interpolate(timeline_left[0], timeline_right[0], ratio)

        points = VGroup()
        stems = VGroup()
        labels = VGroup()
        dates = VGroup()

        grouped = OrderedDict()
        for task in dated:
            key = task["start"].date()
            grouped.setdefault(key, []).append(task)

        above_idx = 0
        below_idx = 0

        above_idx = 0
        below_idx = 0

        for idx, (key, tasks_for_date) in enumerate(grouped.items()):
            x = date_to_x(tasks_for_date[0]["start"])
            y = timeline_left[1]

            point = Dot([x, y, 0], radius=0.065, color=BLUE_D)

            above = idx % 2 == 0
            if above:
                stem_len = [1.0, 1.5, 2.0][above_idx % 3]
                above_idx += 1
            else:
                stem_len = [1.0, 1.5, 2.0][below_idx % 3]
                below_idx += 1

            stem_end_y = y + (stem_len if above else -stem_len)
            stem = Line([x, y, 0], [x, stem_end_y, 0], color=GRAY_B, stroke_width=1)

            date_text = tasks_for_date[0]["start"].strftime("%d/%m")
            date_label = Text(date_text, font_size=12, color=BLUE_D)
            date_label.next_to(point, DOWN if above else UP, buff=0.1)

            offsets = [1.0, 1.5, 2.0] if len(tasks_for_date) > 1 else [stem_len + 0.2]
            for t_idx, task in enumerate(tasks_for_date):
                offset = offsets[t_idx] if t_idx < len(offsets) else offsets[-1]
                target_y = y + (offset if above else -offset)

                title = Text(f"ID {task['id']}", font_size=12, weight=BOLD)
                wrapped_name = textwrap.fill(task["name"], width=28)
                body = Text(wrapped_name, font_size=12, color=GRAY_B, line_spacing=0.9)
                text_block = VGroup(title, body).arrange(DOWN, buff=0.08, aligned_edge=LEFT)

                text_block.move_to([x, target_y, 0])
                labels.add(text_block)

            points.add(point)
            stems.add(stem)
            dates.add(date_label)

        if undated:
            undated_title = Text("Sin fechas", font_size=16, color=GRAY_B)
            undated_lines = VGroup(
                *[Text(f"{t['id']} - {t['name']}", font_size=14) for t in undated]
            ).arrange(DOWN, aligned_edge=LEFT, buff=0.2)
            undated_block = VGroup(undated_title, undated_lines).arrange(DOWN, buff=0.2)
            undated_block.to_edge(RIGHT, buff=0.6).shift(DOWN * 2.2)
        else:
            undated_block = VGroup()

        self.play(Write(header), run_time=1)
        self.play(Create(timeline), run_time=0.8)
        self.play(LaggedStartMap(FadeIn, points, lag_ratio=0.05), run_time=0.9)
        self.play(LaggedStartMap(Create, stems, lag_ratio=0.05), run_time=1.0)
        self.play(LaggedStartMap(FadeIn, dates, lag_ratio=0.05), run_time=0.8)
        self.play(LaggedStartMap(FadeIn, labels, lag_ratio=0.05), run_time=1.2)

        if undated_block:
            self.play(FadeIn(undated_block), run_time=0.6)

        self.wait(2)


class GanttTimelineCircular(ThreeDScene):
    def construct(self):
        source_path = custom_args.source or os.environ.get("GANTT_SOURCE") or DEFAULT_TASK_FILE
        tasks = load_tasks(source_path)
        level2 = [row for row in tasks if row[1] == 2]

        dated = []
        for row in level2:
            task_id, _, name, *_rest, start, end, _pct, _dur, _pred = row
            if start and end:
                dated.append(
                    {
                        "id": task_id,
                        "name": name,
                        "start": datetime.strptime(start, "%d/%m/%y"),
                    }
                )

        dated.sort(key=lambda t: t["start"])

        title = Text("Hablitación Plataforma Calypso Banco BCI", font_size=32, weight=BOLD)
        subtitle = Text("Ambiente Pre Productivo", font_size=18, color=GRAY_B)
        header = VGroup(title, subtitle).arrange(DOWN, buff=0.2).to_edge(UP, buff=0.4)

        outer = Arc(
            radius=3.2,
            start_angle=PI * 0.2,
            angle=PI * 1.6,
            color=GRAY_B,
            stroke_width=2,
        )
        inner = Arc(
            radius=2.6,
            start_angle=PI * 0.2,
            angle=PI * 1.6,
            color=GRAY_C,
            stroke_width=1,
        )

        ticks = VGroup()
        for i in range(70):
            ang = outer.start_angle + (outer.angle * i / 69)
            p1 = 3.25 * np.array([np.cos(ang), np.sin(ang), 0])
            p2 = 3.4 * np.array([np.cos(ang), np.sin(ang), 0])
            ticks.add(Line(p1, p2, stroke_width=1, color=GRAY_B))

        points = VGroup()
        labels = VGroup()
        if dated:
            start_min = min(t["start"] for t in dated)
            end_max = max(t["start"] for t in dated)
        else:
            start_min = datetime.now()
            end_max = datetime.now()

        def date_to_angle(value: datetime) -> float:
            total = (end_max - start_min).days or 1
            offset = (value - start_min).days
            ratio = offset / total
            return outer.start_angle + outer.angle * ratio

        for idx, task in enumerate(dated):
            ang = date_to_angle(task["start"])
            radius = 2.9
            pos = radius * np.array([np.cos(ang), np.sin(ang), 0])
            points.add(Dot(pos, radius=0.05, color=BLUE_D))

            label = Text(f"ID {task['id']}", font_size=12, color=GRAY_B)
            label.move_to((radius + 0.45) * np.array([np.cos(ang), np.sin(ang), 0]))
            label.rotate(ang + PI / 2)
            labels.add(label)

        group = VGroup(outer, inner, ticks, points, labels)
        group.move_to(DOWN * 0.2)
        group.rotate(PI / 5, axis=RIGHT)
        group.rotate(-PI / 10, axis=UP)

        self.set_camera_orientation(phi=70 * DEGREES, theta=-35 * DEGREES)

        self.play(Write(header), run_time=1)
        self.play(Create(outer), Create(inner), run_time=1.2)
        self.play(LaggedStartMap(Create, ticks, lag_ratio=0.02), run_time=1.5)
        self.play(LaggedStartMap(FadeIn, points, lag_ratio=0.05), run_time=1.0)
        self.play(LaggedStartMap(FadeIn, labels, lag_ratio=0.05), run_time=1.0)
        self.play(Rotate(group, angle=PI / 2, axis=UP), run_time=4, rate_func=linear)
        self.wait(1)


if __name__ == "__main__":
    # manim -pql Gantt/Manim/gantt_timeline.py GanttTimelineLevel2
    # manim -pql Gantt/Manim/gantt_timeline.py GanttTimelineCircular
    pass
