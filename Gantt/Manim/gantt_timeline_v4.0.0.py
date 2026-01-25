from __future__ import annotations

import argparse
import ast
import sys
import textwrap
from collections import OrderedDict
from datetime import datetime, timedelta, date
from pathlib import Path
import uuid
import os
import random

from manim import *
from openpyxl import load_workbook


# =============================================================================
# Funciones auxiliares
# =============================================================================
def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%y")
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%y")
        except ValueError:
            continue
    return text


def format_percent(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        pct = value * 100 if 0 <= value <= 1 else value
        return f"{int(round(pct))}%"
    text = str(value).strip()
    if not text:
        return ""
    return text if "%" in text else f"{text}%"


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value


def business_days_in_span(start_date, end_date):
    """Return business days in (start_date, end_date] excluding Sat/Sun."""
    start_d = _as_date(start_date)
    end_d = _as_date(end_date)
    days = []
    total = (end_d - start_d).days
    for i in range(1, total + 1):
        d = start_d + timedelta(days=i)
        if d.weekday() < 5:
            days.append(d)
    return days


HOLIDAYS_2026 = {
    date(2026, 1, 1),
    date(2026, 4, 3),
    date(2026, 4, 4),
    date(2026, 5, 1),
    date(2026, 5, 21),
    date(2026, 6, 21),
    date(2026, 6, 29),
    date(2026, 7, 16),
    date(2026, 8, 15),
    date(2026, 9, 18),
    date(2026, 9, 19),
    date(2026, 10, 12),
    date(2026, 10, 31),
    date(2026, 11, 1),
    date(2026, 12, 8),
    date(2026, 12, 25),
}


def holidays_in_span(start_date, end_date, holidays):
    """Return holidays in (start_date, end_date] that fall on weekdays."""
    start_d = _as_date(start_date)
    end_d = _as_date(end_date)
    days = []
    total = (end_d - start_d).days
    for i in range(1, total + 1):
        d = start_d + timedelta(days=i)
        if d in holidays and d.weekday() < 5:
            days.append(d)
    return days


def business_days_count(start_date, end_date, holidays):
    start_d = _as_date(start_date)
    end_d = _as_date(end_date)
    if end_d < start_d:
        return 0
    count = 0
    cur = start_d
    while cur <= end_d:
        if cur.weekday() < 5 and cur not in holidays:
            count += 1
        cur += timedelta(days=1)
    return count


def load_tasks_from_xlsx(path: Path) -> list[list]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in header_row]
    header_map = {str(val).strip().lower(): idx for idx, val in enumerate(headers, start=1) if val}

    def col_index(*names: str) -> int | None:
        for name in names:
            key = name.strip().lower()
            if key in header_map:
                return header_map[key]
        return None

    col_id = col_index("id", "uid", "uid ")
    col_name = col_index("nombre de la tarea", "nombre", "task name", "name")
    col_status = col_index("estado", "status")
    col_assigned = col_index("asignado a", "asignado", "assigned")
    col_start = col_index("fecha de inicio", "inicio", "start", "start date")
    col_end = col_index("fecha de finalización", "fecha de finalizacion", "fin", "finish", "end", "end date")
    col_pct = col_index("porcentaje completo", "% completo", "avance", "percent complete", "percentcomplete")
    col_duration = col_index("duración", "duracion", "duration")
    col_pred = col_index("predecesores", "predecessors", "pred")

    if not col_name:
        raise ValueError("No se encontro la columna 'Nombre de la tarea' en el XLSX.")

    tasks = []
    for row_idx in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=col_name)
        name_val = name_cell.value
        if name_val is None or str(name_val).strip() == "":
            continue

        indent = name_cell.alignment.indent or 0
        level = int(indent)

        task_id = ws.cell(row=row_idx, column=col_id).value if col_id else row_idx - 1
        status = ws.cell(row=row_idx, column=col_status).value if col_status else ""
        assigned = ws.cell(row=row_idx, column=col_assigned).value if col_assigned else ""
        start = ws.cell(row=row_idx, column=col_start).value if col_start else ""
        end = ws.cell(row=row_idx, column=col_end).value if col_end else ""
        pct = ws.cell(row=row_idx, column=col_pct).value if col_pct else ""
        duration = ws.cell(row=row_idx, column=col_duration).value if col_duration else ""
        pred = ws.cell(row=row_idx, column=col_pred).value if col_pred else ""

        if isinstance(task_id, float) and task_id.is_integer():
            task_id = int(task_id)

        tasks.append(
            [
                task_id,
                level,
                str(name_val).strip(),
                str(status).strip() if status is not None else "",
                str(assigned).strip() if assigned is not None else "",
                format_date(start),
                format_date(end),
                format_percent(pct),
                str(duration).strip() if duration is not None else "",
                str(pred).strip() if pred is not None else "",
            ]
        )

    return tasks


def _parse_levels(values: list[str] | None) -> list[int]:
    if not values:
        return []
    levels: list[int] = []
    for raw in values:
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        for part in parts:
            levels.append(int(part))
    return levels


def filter_by_levels(tasks: list[list], levels: list[int]) -> list[list]:
    """Filtra por niveles e incluye ancestros para dar contexto visual."""
    if not levels:
        return tasks
    level_set = set(levels)
    result: list[list] = []
    seen: set[tuple] = set()
    stack: list[list] = []

    for row in tasks:
        level = row[1]
        while stack and stack[-1][1] >= level:
            stack.pop()
        stack.append(row)

        if level in level_set:
            # Incluir ancestros y la fila actual sin duplicados.
            for parent in stack:
                key = tuple(parent)
                if key not in seen:
                    result.append(parent)
                    seen.add(key)

    return result


def parse_filter_sequence(argv: list[str]) -> list[tuple[str, list[int] | int]]:
    """Parsea --nivel/--id en el orden recibido para aplicar filtros anidados."""
    seq: list[tuple[str, list[int] | int]] = []
    i = 0
    while i < len(argv):
        arg = argv[i]
        if arg == "--nivel":
            if i + 1 >= len(argv):
                raise SystemExit("Error: --nivel requiere un valor.")
            levels = _parse_levels([argv[i + 1]])
            seq.append(("nivel", levels))
            i += 2
            continue
        if arg == "--id":
            if i + 1 >= len(argv):
                raise SystemExit("Error: --id requiere un valor.")
            try:
                task_id = int(argv[i + 1])
            except ValueError as exc:
                raise SystemExit("Error: --id requiere un entero.") from exc
            seq.append(("id", task_id))
            i += 2
            continue
        i += 1
    return seq


def split_by_pipe(argv: list[str]) -> list[list[str]]:
    """Divide argumentos en segmentos separados por '|'."""
    segments: list[list[str]] = []
    current: list[str] = []
    for arg in argv:
        if arg == "|":
            segments.append(current)
            current = []
        else:
            current.append(arg)
    segments.append(current)
    return segments


def filter_by_id_with_context(tasks: list[list], task_id: int, max_depth: int | None = None) -> list[list]:
    """
    Filtra la tarea con el ID dado más sus hijos directos.
    Retorna la tarea padre y todas las tareas con nivel mayor hasta encontrar
    otra tarea del mismo nivel o menor.
    """
    result: list[list] = []
    seen: set[tuple] = set()
    stack: list[list] = []

    found_idx = None
    parent_level = None

    for idx, row in enumerate(tasks):
        level = row[1]
        while stack and stack[-1][1] >= level:
            stack.pop()
        stack.append(row)
        if row[0] == task_id:
            found_idx = idx
            parent_level = level
            for parent in stack:
                key = tuple(parent)
                if key not in seen:
                    result.append(parent)
                    seen.add(key)
            break

    if found_idx is None or parent_level is None:
        return []

    # Agregar hijos (nivel > parent_level) hasta encontrar mismo nivel o menor
    for row in tasks[found_idx + 1:]:
        if row[1] <= parent_level:
            break
        if max_depth is not None and row[1] > parent_level + max_depth:
            continue
        key = tuple(row)
        if key not in seen:
            result.append(row)
            seen.add(key)

    return result


def get_tasks_for_render() -> list[list]:
    """Lee tareas desde filter_gantt.tasks (generado por el CLI)."""
    tasks_file = Path(__file__).with_name("filter_gantt.tasks")
    if not tasks_file.exists():
        raise FileNotFoundError(
            f"No se encontró {tasks_file}. "
            "Ejecuta primero el CLI para generar filter_gantt.tasks."
        )
    return load_tasks_from_file(tasks_file)


def load_tasks_from_file(path: Path) -> list[list]:
    text = path.read_text(encoding="utf-8").lstrip()
    module = ast.parse(text)
    for node in module.body:
        if isinstance(node, ast.Assign):
            for target in node.targets:
                if isinstance(target, ast.Name) and target.id == "tasks":
                    return ast.literal_eval(node.value)
    raise ValueError("tasks list not found in file")


def write_tasks_file(tasks: list[list], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8") as f:
        f.write("tasks = [\n")
        for row in tasks:
            f.write(f"    {row},\n")
        f.write("]\n")


# =============================================================================
# CLI standalone (python gantt_timeline_v2.py -xlsx ... --nivel ...)
# =============================================================================
def run_filter_cli() -> int:
    """CLI para generar filter_gantt.tasks desde XLSX."""
    parser = argparse.ArgumentParser(
        description="Genera filter_gantt.tasks desde XLSX o renderiza con Manim."
    )
    parser.add_argument("-xlsx", "--xlsx", required=True, type=Path, help="Ruta al archivo XLSX.")
    parser.add_argument(
        "--expand",
        action="store_true",
        help="Al usar --id, expande solo el siguiente nivel del ID desde el XLSX completo.",
    )
    parser.add_argument(
        "-debug",
        "--debug",
        action="store_true",
        help="Imprime un informe breve del filtro (IDs/niveles/títulos).",
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=Path(__file__).with_name("filter_gantt.tasks"),
        help="Archivo de salida (default: filter_gantt.tasks).",
    )
    args, _unknown = parser.parse_known_args()
    segments = split_by_pipe(sys.argv[1:])

    if not args.xlsx.exists():
        print(f"Error: no existe el archivo {args.xlsx}", file=sys.stderr)
        return 1

    filtered = load_tasks_from_xlsx(args.xlsx)
    temp_files: list[Path] = []

    for seg_idx, seg in enumerate(segments):
        filter_seq = parse_filter_sequence(seg)
        if filter_seq:
            for kind, value in filter_seq:
                if kind == "nivel":
                    filtered = filter_by_levels(
                        filtered, value if isinstance(value, list) else [value]
                    )
                    print(f"Filtrado por nivel {value}: {len(filtered)} tareas")
                elif kind == "id":
                    base = load_tasks_from_xlsx(args.xlsx) if args.expand else filtered
                    depth = 1 if args.expand else None
                    filtered = filter_by_id_with_context(base, int(value), max_depth=depth)
                    print(f"Filtrado por ID {value}: {len(filtered)} tareas")
        else:
            if seg_idx == 0:
                print(f"Sin filtro: {len(filtered)} tareas")

        if seg_idx < len(segments) - 1:
            tmp = Path("/tmp") / f"gantt_filter_{uuid.uuid4().hex}_{seg_idx}.tasks"
            write_tasks_file(filtered, tmp)
            temp_files.append(tmp)
            filtered = load_tasks_from_file(tmp)

    if args.debug:
        levels = sorted({row[1] for row in filtered})
        print(f"Niveles presentes: {levels}")
        print("Tareas filtradas (id | nivel | título):")
        for task_id, level, name, *_rest in filtered:
            print(f"- {task_id} | {level} | {name}")

    write_tasks_file(filtered, args.output)
    print(f"Escrito: {args.output}")

    for tmp in temp_files:
        try:
            os.remove(tmp)
        except OSError:
            pass
    return 0


# =============================================================================
# Escenas Manim
# =============================================================================

# Uso:
#   manim -pql gantt_timeline_v2.py GanttTimelineLevel2 --xlsx archivo.xlsx --nivel 1
#   manim -pql gantt_timeline_v2.py GanttTimelineLevel2 --xlsx archivo.xlsx --id 42


class GanttTimelineLevel2(Scene):
    def construct(self):
        tasks = get_tasks_for_render()

        title_text = "Hablitación Plataforma Calypso Banco BCI"
        subtitle_text = "Ambiente Pre Productivo"
        level0 = next((row for row in tasks if row[1] == 0), None)
        level1 = next((row for row in tasks if row[1] == 1), None)
        if level0:
            title_text = level0[2]
        if level1:
            subtitle_text = level1[2]

        tasks = [row for row in tasks if row[1] >= 2]

        dated = []
        undated = []
        for row in tasks:
            task_id, _, name, *_rest, start, end, pct, _dur, _pred = row
            if start and end:
                dated.append(
                    {
                        "id": task_id,
                        "name": name,
                        "start": datetime.strptime(start, "%d/%m/%y"),
                        "end": datetime.strptime(end, "%d/%m/%y"),
                        "start_str": start,
                        "end_str": end,
                        "pct": pct,
                    }
                )
            else:
                undated.append({"id": task_id, "name": name})

        dated.sort(key=lambda t: t["start"])

        title = Text(title_text, font_size=28, weight=BOLD)
        subtitle = Text(subtitle_text, font_size=16, color=GRAY_B)
        header = VGroup(title, subtitle).arrange(DOWN, buff=0.2).to_corner(UL, buff=0.4)

        # Contador estilo "flip" con fecha (animable)
        today_dt = datetime.now()
        start_date = date(today_dt.year, 1, 6)
        current_date = start_date
        counter_labels = ["DIA", "MES", "ANO", "ANO"]
        counter_values = [
            f"{current_date.day:02d}",
            f"{current_date.month:02d}",
            f"{current_date.year // 100:02d}",
            f"{current_date.year % 100:02d}",
        ]
        counter_boxes = VGroup()
        counter_blocks: list[dict[str, object]] = []
        for label, value in zip(counter_labels, counter_values):
            box = RoundedRectangle(
                width=0.76,
                height=0.48,
                corner_radius=0.06,
                stroke_width=1,
                stroke_color=GRAY_D,
                fill_color=BLACK,
                fill_opacity=0.35,
            )
            # División central tipo reloj de aeropuerto (prepara animación de "flip")
            mid_y = box.get_center()[1]
            split_line = Line(
                [box.get_left()[0], mid_y, 0],
                [box.get_right()[0], mid_y, 0],
                color=GRAY_C,
                stroke_width=1,
                stroke_opacity=0.6,
            )
            # Mitad superior/inferior con leve contraste
            top_mask = Rectangle(
                width=box.width,
                height=box.height / 2,
                stroke_width=0,
                fill_color=BLACK,
                fill_opacity=0.22,
            ).move_to([box.get_center()[0], mid_y + box.height / 4, 0])
            bottom_mask = Rectangle(
                width=box.width,
                height=box.height / 2,
                stroke_width=0,
                fill_color=BLACK,
                fill_opacity=0.38,
            ).move_to([box.get_center()[0], mid_y - box.height / 4, 0])

            value_text = Text(value, font_size=18, weight=BOLD, color=WHITE)
            value_text.move_to(box.get_center() + DOWN * 0.03)
            label_text = Text(label, font_size=7, color=GRAY_B)
            label_text.next_to(box, UP, buff=0.06)
            card = VGroup(box, top_mask, bottom_mask, split_line, value_text)
            group = VGroup(label_text, card)
            counter_boxes.add(group)
            counter_blocks.append({"group": group, "card": card})
        counter_boxes.arrange(RIGHT, buff=0.18, aligned_edge=DOWN)
        counter_boxes.to_corner(UR, buff=0.46)
        counter_boxes.shift(UP * 0.04)

        timeline_left = LEFT * 5.5 + DOWN * 0.2
        timeline_right = RIGHT * 5.5 + DOWN * 0.2
        timeline = Line(timeline_left, timeline_right, color=GRAY_B, stroke_width=4)

        if dated:
            start_min = min(t["start"] for t in dated)
            end_max = max(t["end"] for t in dated)
        else:
            start_min = datetime.now()
            end_max = datetime.now()

        def date_to_x(value: datetime) -> float:
            total = (end_max - start_min).days or 1
            offset = (value - start_min).days
            ratio = offset / total
            return interpolate(timeline_left[0], timeline_right[0], ratio)

        scale_y = timeline_left[1] - 3.45
        tlu_label = Text("TLU", font_size=12, color=GRAY_B)
        tlu_label.next_to(timeline_left, LEFT, buff=0.4)
        tmd_label = Text("TLD", font_size=12, color=GRAY_B)
        tmd_label.next_to([timeline_left[0], scale_y, 0], LEFT, buff=0.4)

        # Fecha de hoy (ajuste de año solo si cae dentro del rango del Gantt)
        today = datetime.now()
        if today.year != start_min.year:
            try:
                candidate = today.replace(year=start_min.year)
            except ValueError:
                candidate = today.replace(year=start_min.year, day=28)
            if start_min <= candidate <= end_max:
                today = candidate

        # Promedio global de avance real y planificado para marcador de "hoy"
        pct_all = []
        for t in dated:
            if t.get("pct"):
                try:
                    pct_all.append(float(str(t["pct"]).replace("%", "").strip()))
                except ValueError:
                    pass
        avg_all = round(sum(pct_all) / len(pct_all)) if pct_all else None
        planned_all = []
        for t in dated:
            start_d = t["start"]
            end_d = t["end"]
            total_days = max(1, (end_d - start_d).days)
            if today <= start_d:
                planned = 0.0
            elif today >= end_d:
                planned = 100.0
            else:
                elapsed = (today - start_d).days
                planned = (elapsed / total_days) * 100.0
            planned_all.append(planned)
        avg_planned = round(sum(planned_all) / len(planned_all)) if planned_all else None

        # Línea de "hoy" interpolada entre puntos vecinos para respetar separaciones reales
        date_keys = [t["start"].date() for t in dated]
        if date_keys:
            today_date = today.date()
            prev_d = max((d for d in date_keys if d <= today_date), default=date_keys[0])
            next_d = min((d for d in date_keys if d >= today_date), default=date_keys[-1])
            x_prev = date_to_x(datetime.combine(prev_d, datetime.min.time()))
            x_next = date_to_x(datetime.combine(next_d, datetime.min.time()))
            span = (next_d - prev_d).days or 1
            ratio_local = (today_date - prev_d).days / span
            x_today = x_prev + (x_next - x_prev) * ratio_local
        else:
            x_today = date_to_x(today)
        # Dial vintage: barras paralelas para real vs plan (más separación si difieren)
        dial_height = 0.55
        dial_w = 0.08
        plan_val = avg_planned if avg_planned is not None else 0
        real_val = avg_all if avg_all is not None else 0
        diff = abs((real_val or 0) - (plan_val or 0))
        dial_gap = 0.02 + 0.0015 * diff
        dial_gap = min(0.22, dial_gap)
        dial_real = Rectangle(
            width=dial_w,
            height=dial_height,
            stroke_width=0,
            fill_color=GREEN_E,
            fill_opacity=0.35,
        ).move_to([x_today - dial_gap / 2, scale_y + dial_height / 2, 0])
        dial_plan = Rectangle(
            width=dial_w,
            height=dial_height,
            stroke_width=0,
            fill_color=GREEN_A,
            fill_opacity=0.35,
        ).move_to([x_today + dial_gap / 2, scale_y + dial_height / 2, 0])
        today_line = VGroup(dial_real, dial_plan)
        today_tick = Line(
            [x_today, timeline_left[1] - 0.18, 0],
            [x_today, timeline_left[1] + 0.18, 0],
            color=GREEN_E,
            stroke_width=1,
        )
        today_label = Text(f"HOY {today.strftime('%d/%m')}", font_size=11, color=GREEN_E)
        pct_parts = []
        if avg_all is not None:
            pct_parts.append(f"R {avg_all}%")
        if avg_planned is not None:
            pct_parts.append(f"P {avg_planned}%")
        if pct_parts:
            today_pct = Text(" | ".join(pct_parts), font_size=8, color=GREEN_E)
        else:
            today_pct = None
        total_days = business_days_count(start_min, end_max, HOLIDAYS_2026)
        calendar_days = (end_max.date() - start_min.date()).days + 1
        holidays_span = holidays_in_span(start_min, end_max, HOLIDAYS_2026)
        holidays_count = len(holidays_span)
        elapsed_end = min(today.date(), end_max.date())
        elapsed_days = business_days_count(start_min, elapsed_end, HOLIDAYS_2026)
        elapsed_pct = int(round((elapsed_days / total_days) * 100)) if total_days else 0
        today_days = Text(f"HAB {total_days}d", font_size=8, color=GREEN_E)
        today_elapsed = Text(f"TRANS {elapsed_days}d", font_size=8, color=GREEN_E)
        today_elapsed_pct = Text(f"{elapsed_pct}%", font_size=8, color=GREEN_E)
        today_info = VGroup(today_label, today_pct, today_days, today_elapsed, today_elapsed_pct).arrange(
            DOWN, buff=0.04, aligned_edge=LEFT
        )
        today_info.move_to([timeline_left[0] - 1.0, scale_y + dial_height / 2 + 0.12, 0])
        today_info_line = VGroup()
        line_y = scale_y + dial_height / 2 + 0.12
        line_left = x_today - 0.12
        line_right = today_info.get_right()[0]
        line_segs = 10
        min_opacity = 0.1
        for s in range(line_segs):
            t0 = s / line_segs
            t1 = (s + 1) / line_segs
            x0 = line_left + (line_right - line_left) * t0
            x1 = line_left + (line_right - line_left) * t1
            t_mid = (t0 + t1) / 2
            opacity = min_opacity + (1 - min_opacity) * abs(2 * t_mid - 1)
            seg = Line([x0, line_y, 0], [x1, line_y, 0], color=GREEN_E, stroke_width=1, stroke_opacity=opacity)
            today_info_line.add(seg)
        # punta simple: tres puntos decrecientes hacia la barra (lado de x_today)
        tip_radii = [0.04, 0.028, 0.018]
        tip_offsets = [-0.16, -0.07, 0.0]
        for r, dx in zip(tip_radii, tip_offsets):
            tip = Dot([line_left + dx, line_y, 0], radius=r, color=GREEN_E)
            today_info_line.add(tip)

        if "DEBUG_TODAY" in os.environ:
            print(f"[DEBUG_TODAY] start_min={start_min.date()} end_max={end_max.date()} today={today.date()}")

        points = VGroup()
        end_points = VGroup()
        end_dates = VGroup()
        connectors = VGroup()
        connector_ends = VGroup()
        stems_bg = VGroup()
        stems_lit = VGroup()
        labels = VGroup()
        dates = VGroup()
        deltas = VGroup()
        full_test = VGroup()
        full_test_segments = VGroup()
        pct_by_date: dict = {}

        grouped = OrderedDict()
        for task in dated:
            key = task["start"].date()
            grouped.setdefault(key, []).append(task)

        end_grouped = OrderedDict()
        for task in dated:
            if task["end"].date() == task["start"].date():
                continue
            end_key = task["end"].date()
            end_grouped.setdefault(end_key, []).append(task)

        above_idx = 0
        below_idx = 0
        spacing_scale = 0.85

        date_keys = list(grouped.keys())
        for idx, (key, tasks_for_date) in enumerate(grouped.items()):
            x = date_to_x(tasks_for_date[0]["start"])
            y = timeline_left[1]

            point = Dot([x, y, 0], radius=0.065, color=RED_E)

            above = idx % 2 == 0
            if above:
                stem_len = [1.0, 1.5, 2.0][above_idx % 3] * spacing_scale
                above_idx += 1
            else:
                stem_len = [1.0, 1.5, 2.0][below_idx % 3] * spacing_scale
                below_idx += 1

            date_text = tasks_for_date[0]["start"].strftime("%d/%m")
            date_label = Text(date_text, font_size=12, color=RED_E)
            date_label.next_to(point, DOWN if above else UP, buff=0.1)

            offsets = [1.0, 1.5, 2.0]
            if len(tasks_for_date) > 1:
                offsets = [o * spacing_scale for o in offsets]
            else:
                offsets = [stem_len + 0.2 * spacing_scale]
            x_offsets = [0.05, -0.05, 0.12, -0.12, 0.2, -0.2]
            for t_idx, task in enumerate(tasks_for_date):
                offset = offsets[t_idx] if t_idx < len(offsets) else offsets[-1]
                target_y = y + (offset if above else -offset)

                id_line = f"ID {task['id']}"
                if task.get("pct"):
                    id_line = f"{id_line}  {task['pct']}"
                title_text = Text(id_line, font_size=12, weight=BOLD)
                end_text = Text(f"Fin: {task['end_str']}", font_size=11, color=GRAY_C)
                text_block = VGroup(title_text, end_text).arrange(DOWN, buff=0.06, aligned_edge=LEFT)

                text_block.move_to([x, target_y, 0])
                x_shift = x_offsets[t_idx] if t_idx < len(x_offsets) else x_offsets[-1]
                text_block.shift(RIGHT * x_shift)
                if target_y >= y:
                    text_block.shift(UP * 0.15)
                else:
                    text_block.shift(DOWN * 0.15)
                labels.add(text_block)

                # Barra vertical tipo ecualizador (segmentos horizontales)
                bar_width = 0.18
                bar_height = abs(target_y - y) * 0.9
                bar_center = (y + target_y) / 2

                pct_val = None
                if task.get("pct"):
                    try:
                        pct_val = float(str(task["pct"]).replace("%", "").strip())
                    except ValueError:
                        pct_val = None
                if pct_val is None:
                    stem = Line(
                        [x, y, 0],
                        [x, target_y, 0],
                        color=GRAY_C,
                        stroke_width=2,
                    )
                    stems_bg.add(stem)
                    continue
                # Siempre dibujar segmentos de fondo (apagados)
                seg_count = 14
                gap_ratio = 0.2
                seg_height = bar_height / (seg_count + (seg_count - 1) * gap_ratio)
                seg_gap = seg_height * gap_ratio
                fill_bg = VGroup()
                fill_lit = VGroup()
                halo = Rectangle(
                    width=bar_width,
                    height=bar_height,
                    stroke_width=0,
                    fill_color=GRAY_E,
                    fill_opacity=0.12,
                ).move_to([x, bar_center, 0])
                fill_bg.add(halo)
                for s in range(seg_count):
                    seg = Rectangle(
                        width=bar_width,
                        height=seg_height,
                        stroke_width=0,
                        fill_color=GRAY_C,
                        fill_opacity=0.28,
                    )
                    if target_y >= y:
                        seg_y = y + (seg_height / 2) + s * (seg_height + seg_gap)
                    else:
                        seg_y = y - (seg_height / 2) - s * (seg_height + seg_gap)
                    seg.move_to([x, seg_y, 0])
                    fill_bg.add(seg)

                # Encender segmentos según % (si hay valor)
                if pct_val is not None and pct_val > 0:
                    pct_norm = max(0.0, min(1.0, pct_val / 100.0))
                    lit_segments = int(round(pct_norm * seg_count))
                    if pct_norm > 0 and lit_segments == 0:
                        lit_segments = 1
                    for s in range(lit_segments):
                        t = s / max(1, seg_count - 1)
                        if t <= 0.5:
                            color = interpolate_color(RED_E, GREEN_B, t * 2)
                        else:
                            color = interpolate_color(GREEN_B, BLUE_E, (t - 0.5) * 2)
                        seg = Rectangle(
                            width=bar_width,
                            height=seg_height,
                            stroke_width=0,
                            fill_color=color,
                            fill_opacity=0.95,
                        )
                        if target_y >= y:
                            seg_y = y + (seg_height / 2) + s * (seg_height + seg_gap)
                        else:
                            seg_y = y - (seg_height / 2) - s * (seg_height + seg_gap)
                        seg.move_to([x, seg_y, 0])
                        fill_lit.add(seg)

                # Overlay para prueba de calidad: 100% lleno (se anima al final)
                full_fill = VGroup()
                for s in range(seg_count):
                    t = s / max(1, seg_count - 1)
                    if t <= 0.5:
                        color = interpolate_color(RED_E, GREEN_B, t * 2)
                    else:
                        color = interpolate_color(GREEN_B, BLUE_E, (t - 0.5) * 2)
                    seg = Rectangle(
                        width=bar_width,
                        height=seg_height,
                        stroke_width=0,
                        fill_color=color,
                        fill_opacity=0.8,
                    )
                    if target_y >= y:
                        seg_y = y + (seg_height / 2) + s * (seg_height + seg_gap)
                    else:
                        seg_y = y - (seg_height / 2) - s * (seg_height + seg_gap)
                    seg.move_to([x, seg_y, 0])
                    full_fill.add(seg)
                    full_test_segments.add(seg)
                full_test.add(full_fill)
                # Cap para 100%: asegura que llegue al borde
                if pct_val is not None and pct_val >= 99.9:
                    cap_h = seg_height * 0.6
                    cap = Rectangle(
                        width=bar_width,
                        height=cap_h,
                        stroke_width=0,
                        fill_color=YELLOW_B,
                        fill_opacity=0.95,
                    )
                    if target_y >= y:
                        cap.move_to([x, y + bar_height - cap_h / 2, 0])
                    else:
                        cap.move_to([x, y - bar_height + cap_h / 2, 0])
                    fill_lit.add(cap)

                stems_bg.add(fill_bg)
                if len(fill_lit) > 0:
                    stems_lit.add(fill_lit)

            # Marcas de escala (0-100) junto a la barra (solo una vez por fecha)
            scale_marks = VGroup()
            ticks = [0, 25, 50, 75, 100]
            for t in ticks:
                frac = t / 100.0
                tick_len = 0.16 if t in (0, 50, 100) else 0.08
                if above:
                    ty = y + frac * stem_len
                else:
                    ty = y - frac * stem_len
                scale_marks.add(
                    Line([x - 0.18, ty, 0], [x - 0.18 - tick_len, ty, 0], color=GRAY_C, stroke_width=1)
                )
                if t in (0, 50, 100):
                    lbl = Text(str(t), font_size=9, color=GRAY_C)
                    lbl.next_to(scale_marks[-1], LEFT, buff=0.04)
                    scale_marks.add(lbl)

            points.add(point)
            dates.add(date_label)
            stems_bg.add(scale_marks)

            pcts = []
            for t in tasks_for_date:
                if t.get("pct"):
                    try:
                        pcts.append(float(str(t["pct"]).replace("%", "").strip()))
                    except ValueError:
                        pass
            if pcts:
                pct_by_date[key] = round(sum(pcts) / len(pcts))

        # Marcar fechas de fin en la escala inferior (solo punto + fecha)
        end_keys_sorted = sorted(end_grouped.keys())
        for idx, end_key in enumerate(end_keys_sorted):
            x_end = date_to_x(datetime.combine(end_key, datetime.min.time()))
            y = scale_y
            end_point = Dot([x_end, y, 0], radius=0.05, color=BLUE_D)
            end_label = Text(end_key.strftime("%d/%m"), font_size=11, color=BLUE_D)
            if idx % 2 == 0:
                end_label.next_to(end_point, DOWN, buff=0.08)
            else:
                end_label.next_to(end_point, UP, buff=0.08)
            end_points.add(end_point)
            end_dates.add(end_label)

        # Escala inferior estilo "mapa": barra segmentada con dias por tramo
        bar_height = 0.15
        bar_bg = VGroup()
        bar_lit = VGroup()
        bar_full = VGroup()
        date_guides = VGroup()
        holiday_marks = VGroup()
        scale_keys = sorted(set(date_keys + [t["end"].date() for t in dated]))
        for i in range(1, len(scale_keys)):
            d0 = scale_keys[i - 1]
            d1 = scale_keys[i]
            business_days = business_days_in_span(d0, d1)
            biz_count = len(business_days)
            holiday_days = holidays_in_span(d0, d1, HOLIDAYS_2026)
            holiday_set = set(holiday_days)
            holiday_count = len(holiday_days)
            x0 = date_to_x(datetime.combine(d0, datetime.min.time()))
            x1 = date_to_x(datetime.combine(d1, datetime.min.time()))
            mid_x = (x0 + x1) / 2

            seg_width = max(0.01, x1 - x0)
            days_for_layout = max(1, biz_count)
            unit_w = seg_width / days_for_layout
            # Usa el % de la fecha de inicio del tramo para evitar corrimientos
            pct_for_span = pct_by_date.get(d0)
            pct_norm = 0.0 if pct_for_span is None else max(0.0, min(1.0, pct_for_span / 100.0))
            if biz_count == 0:
                t_prog = 0.0
                # Fondo apagado
                bg_seg = Rectangle(
                    width=unit_w * 0.85,
                    height=bar_height,
                    stroke_width=0,
                    fill_color=GRAY_C,
                    fill_opacity=0.32,
                )
                bg_seg.move_to([x0 + 0.5 * unit_w, scale_y, 0])
                bar_bg.add(bg_seg)

                full_seg = Rectangle(
                    width=unit_w * 0.85,
                    height=bar_height,
                    stroke_width=0,
                    fill_color=interpolate_color(RED_E, GREEN_B, t_prog * 2),
                    fill_opacity=1,
                )
                full_seg.move_to([x0 + 0.5 * unit_w, scale_y, 0])
                bar_full.add(full_seg)
            else:
                for idx_day, day in enumerate(business_days):
                    seg_x = x0 + (idx_day + 0.5) * unit_w
                    if day in holiday_set:
                        holiday_label = Text(day.strftime("%d/%m"), font_size=9, color=RED_E)
                        if idx_day % 2 == 0:
                            holiday_label.next_to([seg_x, scale_y, 0], DOWN, buff=0.08)
                        else:
                            holiday_label.next_to([seg_x, scale_y, 0], UP, buff=0.08)
                        holiday_marks.add(holiday_label)
                        continue

                    t_prog = idx_day / max(1, biz_count - 1)
                    # Fondo apagado
                    bg_seg = Rectangle(
                        width=unit_w * 0.85,
                        height=bar_height,
                        stroke_width=0,
                        fill_color=GRAY_C,
                        fill_opacity=0.32,
                    )
                    bg_seg.move_to([seg_x, scale_y, 0])
                    bar_bg.add(bg_seg)

                    # Full test (100%)
                    if t_prog <= 0.5:
                        full_color = interpolate_color(RED_E, GREEN_B, t_prog * 2)
                    else:
                        full_color = interpolate_color(GREEN_B, BLUE_E, (t_prog - 0.5) * 2)
                    full_seg = Rectangle(
                        width=unit_w * 0.85,
                        height=bar_height,
                        stroke_width=0,
                        fill_color=full_color,
                        fill_opacity=1,
                    )
                    full_seg.move_to([seg_x, scale_y, 0])
                    bar_full.add(full_seg)

                    if pct_norm > 0 and t_prog <= pct_norm:
                        if t_prog <= 0.5:
                            color = interpolate_color(RED_E, GREEN_B, t_prog * 2)
                        else:
                            color = interpolate_color(GREEN_B, BLUE_E, (t_prog - 0.5) * 2)
                        opacity = 1
                    else:
                        color = None
                        opacity = 0
                    seg = Rectangle(
                        width=unit_w * 0.85,
                        height=bar_height,
                        stroke_width=0,
                        fill_color=color if color else GRAY_C,
                        fill_opacity=opacity,
                    )
                    seg.move_to([seg_x, scale_y, 0])
                    if opacity > 0:
                        bar_lit.add(seg)

            tick = Line([x0, scale_y + 0.08, 0], [x0, scale_y - 0.08, 0], color=GRAY_B, stroke_width=1)
            if holiday_count > 0:
                base = Text(f"{biz_count}d", font_size=9, color=GRAY_B)
                minus = Text(f"-{holiday_count}", font_size=9, color=RED_E)
                label_group = VGroup(base, minus).arrange(RIGHT, buff=0.02)
                label_group.move_to([mid_x, scale_y - 0.28, 0])
                deltas.add(VGroup(tick, label_group))
            else:
                txt = Text(f"{biz_count}d", font_size=9, color=GRAY_B)
                txt.move_to([mid_x, scale_y - 0.28, 0])
                deltas.add(VGroup(tick, txt))


        if scale_keys:
            x_start = date_to_x(datetime.combine(scale_keys[0], datetime.min.time()))
            x_end = date_to_x(datetime.combine(scale_keys[-1], datetime.min.time()))
            deltas.add(Line([x_end, scale_y + 0.08, 0], [x_end, scale_y - 0.08, 0], color=GRAY_B, stroke_width=1))
            # Guías finas desde la fecha superior hacia la escala inferior
            for idx, d in enumerate(date_keys):
                x = date_to_x(datetime.combine(d, datetime.min.time()))
                # Guía con desvanecido en el centro
                y_top = timeline_left[1]
                y_bottom = scale_y - 0.28
                segs = 7
                min_opacity = 0.1
                for s in range(segs):
                    t0 = s / segs
                    t1 = (s + 1) / segs
                    y0 = y_top + (y_bottom - y_top) * t0
                    y1 = y_top + (y_bottom - y_top) * t1
                    t_mid = (t0 + t1) / 2
                    opacity = min_opacity + (1 - min_opacity) * abs(2 * t_mid - 1)
                    seg = Line([x, y0, 0], [x, y1, 0], color=GRAY_B, stroke_width=0.5, stroke_opacity=opacity)
                    date_guides.add(seg)

        # Conectores inicio-fin en TMD (lineas horizontales con desvanecido)
        def _top_or_scale(group):
            return group.get_top()[1] if len(group) > 0 else scale_y

        base_y = max(
            scale_y + 0.18,
            _top_or_scale(end_dates) + 0.18,
            _top_or_scale(holiday_marks) + 0.18,
        )
        levels_count = max(4, min(10, len(dated)))
        level_step = 0.18
        connector_levels = [base_y + i * level_step for i in range(levels_count)]

        for idx, task in enumerate(dated):
            if task["end"].date() == task["start"].date():
                continue
            rng = random.Random(task["id"])
            x_start = date_to_x(task["start"])
            x_end = date_to_x(task["end"])
            if x_end < x_start:
                x_start, x_end = x_end, x_start
            y = connector_levels[idx % len(connector_levels)]
            segs = 10
            min_opacity = 0.1
            for s in range(segs):
                t0 = s / segs
                t1 = (s + 1) / segs
                x0 = x_start + (x_end - x_start) * t0
                x1 = x_start + (x_end - x_start) * t1
                t_mid = (t0 + t1) / 2
                opacity = min_opacity + (1 - min_opacity) * abs(2 * t_mid - 1)
                seg = Line([x0, y, 0], [x1, y, 0], color=GRAY_B, stroke_width=0.6, stroke_opacity=opacity)
                connectors.add(seg)
            # Marca de inicio: punto en el extremo y bajada suave hasta TMD
            def star_burst(cx, cy, color, jitter):
                star = VGroup()
                base_r = 0.034 + jitter.uniform(-0.004, 0.004)
                steps = 5
                for direction in [(1, 0), (-1, 0), (0, 1), (0, -1)]:
                    dx, dy = direction
                    for i in range(1, steps + 1):
                        t = i / steps
                        radius = base_r * (0.55 - 0.35 * t)
                        opacity = 0.75 - 0.6 * t
                        offset = base_r * 2.4 * t
                        x = cx + dx * offset
                        y = cy + dy * offset
                        star.add(
                            Dot([x, y, 0], radius=max(0.006, radius), color=color).set_opacity(opacity)
                        )
                return star

            start_blob = star_burst(x_start, y, RED_E, rng)
            drop_segs = 10
            for s in range(drop_segs):
                t0 = s / drop_segs
                t1 = (s + 1) / drop_segs
                y0 = y + (scale_y - y) * t0
                y1 = y + (scale_y - y) * t1
                t_mid = (t0 + t1) / 2
                opacity = min_opacity + (1 - min_opacity) * abs(2 * t_mid - 1)
                vseg = Line([x_start, y0, 0], [x_start, y1, 0], color=RED_E, stroke_width=0.6, stroke_opacity=opacity)
                connector_ends.add(vseg)
            connector_ends.add(start_blob)
            # Marca de fin: punto en el extremo y bajada suave hasta TMD
            end_blob = star_burst(x_end, y, BLUE_D, rng)
            drop_segs = 10
            for s in range(drop_segs):
                t0 = s / drop_segs
                t1 = (s + 1) / drop_segs
                y0 = y + (scale_y - y) * t0
                y1 = y + (scale_y - y) * t1
                t_mid = (t0 + t1) / 2
                opacity = min_opacity + (1 - min_opacity) * abs(2 * t_mid - 1)
                vseg = Line([x_end, y0, 0], [x_end, y1, 0], color=BLUE_D, stroke_width=0.6, stroke_opacity=opacity)
                connector_ends.add(vseg)
            connector_ends.add(end_blob)

        if undated:
            undated_title = Text("Sin fechas", font_size=16, color=GRAY_B)
            undated_lines = VGroup(
                *[Text(f"{t['id']} - {t['name']}", font_size=14) for t in undated]
            ).arrange(DOWN, aligned_edge=LEFT, buff=0.2)
            undated_block = VGroup(undated_title, undated_lines).arrange(DOWN, buff=0.2)
            undated_block.to_edge(RIGHT, buff=0.6).shift(DOWN * 2.2)
        else:
            undated_block = VGroup()

        self.play(Write(header), run_time=1)
        self.play(FadeIn(counter_boxes), run_time=0.6)

        # Animacion: avanzar desde 06/01 hasta hoy, solo por dias (con linea verde)
        def _flip_value(
            block: dict[str, object],
            new_value: str,
            run_time: float,
            extra_anims: list[Animation] | None = None,
        ) -> None:
            old_card: VGroup = block["card"]  # type: ignore[assignment]
            drop = 0.16
            new_text = Text(new_value, font_size=18, weight=BOLD, color=WHITE)
            new_text.move_to(old_card[4].get_center())
            new_card = VGroup(
                old_card[0].copy(),
                old_card[1].copy(),
                old_card[2].copy(),
                old_card[3].copy(),
                new_text,
            )
            new_card.move_to(old_card.get_center() + UP * drop)
            new_card.set_opacity(0)
            self.add(new_card)
            anims = [
                old_card.animate.shift(DOWN * drop).set_opacity(0),
                new_card.animate.shift(DOWN * drop).set_opacity(1),
            ]
            if extra_anims:
                anims.extend(extra_anims)
            self.play(*anims, run_time=run_time)
            block["group"].remove(old_card)  # type: ignore[call-arg]
            self.remove(old_card)
            block["group"].add(new_card)  # type: ignore[call-arg]
            block["card"] = new_card
        self.play(Create(timeline), run_time=0.8)
        self.play(FadeIn(tlu_label), FadeIn(tmd_label), run_time=0.4)
        today_group = None
        if today_line:
            if today_pct:
                today_group = VGroup(today_line, today_tick, today_info_line, today_info)
            else:
                today_group = VGroup(today_line, today_tick, today_info_line, today_info)
        self.play(LaggedStartMap(FadeIn, points, lag_ratio=0.05), run_time=0.9)
        self.play(LaggedStartMap(FadeIn, stems_bg, lag_ratio=0.05), run_time=1.0)
        self.play(LaggedStartMap(FadeIn, dates, lag_ratio=0.05), run_time=0.8)
        if connectors:
            self.play(LaggedStartMap(FadeIn, connectors, lag_ratio=0.01), run_time=0.6)
        if connector_ends:
            self.play(LaggedStartMap(FadeIn, connector_ends, lag_ratio=0.02), run_time=0.5)
        if end_points:
            self.play(LaggedStartMap(FadeIn, end_points, lag_ratio=0.05), run_time=0.5)
            self.play(LaggedStartMap(FadeIn, end_dates, lag_ratio=0.05), run_time=0.5)
        self.play(LaggedStartMap(FadeIn, labels, lag_ratio=0.05), run_time=1.2)
        if deltas:
            self.play(FadeIn(bar_bg), run_time=0.4)
            self.play(LaggedStartMap(FadeIn, deltas, lag_ratio=0.03), run_time=0.6)
        if date_guides:
            self.play(LaggedStartMap(FadeIn, date_guides, lag_ratio=0.02), run_time=0.4)
        if holiday_marks:
            self.play(LaggedStartMap(FadeIn, holiday_marks, lag_ratio=0.05), run_time=0.5)

        # Mostrar "hoy" junto con el resto de elementos
        if today_group:
            self.play(FadeIn(today_group), run_time=0.4)

        # Prueba de calidad: flash rápido sin pausas perceptibles
        if full_test_segments:
            self.add(full_test, bar_full)
            self.wait(0.5)
            self.remove(full_test, bar_full)

        # Mostrar valores reales después de la prueba
        if stems_lit:
            self.play(LaggedStartMap(FadeIn, stems_lit, lag_ratio=0.05), run_time=0.8)
        if bar_lit:
            self.play(FadeIn(bar_lit), run_time=0.4)

        if undated_block:
            self.play(FadeIn(undated_block), run_time=0.6)

        # Linea de tiempo verde + avance sincronizado con el reloj
        x_start = date_to_x(datetime.combine(start_date, datetime.min.time()))
        progress_line = Line(
            [x_start, timeline_left[1], 0],
            [x_start, timeline_left[1], 0],
            color=GREEN_E,
            stroke_width=4,
        )
        self.play(FadeIn(progress_line), run_time=0.3)

        days_to_advance = (today_dt.date() - start_date).days
        flip_time = 0.08
        if days_to_advance > 0:
            for _ in range(days_to_advance):
                next_date = current_date + timedelta(days=1)
                new_x = date_to_x(datetime.combine(next_date, datetime.min.time()))
                line_anim = progress_line.animate.put_start_and_end_on(
                    [x_start, timeline_left[1], 0],
                    [new_x, timeline_left[1], 0],
                )
                _flip_value(counter_blocks[0], f"{next_date.day:02d}", flip_time, [line_anim])
                if next_date.month != current_date.month:
                    _flip_value(counter_blocks[1], f"{next_date.month:02d}", flip_time)
                if next_date.year != current_date.year:
                    _flip_value(counter_blocks[2], f"{next_date.year // 100:02d}", flip_time)
                    _flip_value(counter_blocks[3], f"{next_date.year % 100:02d}", flip_time)
                current_date = next_date

        self.wait(2)


class GanttTimelineCircular(ThreeDScene):
    def construct(self):
        tasks = get_tasks_for_render()

        title_text = "Hablitación Plataforma Calypso Banco BCI"
        subtitle_text = "Ambiente Pre Productivo"
        level0 = next((row for row in tasks if row[1] == 0), None)
        level1 = next((row for row in tasks if row[1] == 1), None)
        if level0:
            title_text = level0[2]
        if level1:
            subtitle_text = level1[2]

        tasks = [row for row in tasks if row[1] >= 2]

        dated = []
        for row in tasks:
            task_id, _, name, *_rest, start, end, _pct, _dur, _pred = row
            if start and end:
                dated.append(
                    {
                        "id": task_id,
                        "name": name,
                        "start": datetime.strptime(start, "%d/%m/%y"),
                    }
                )

        dated.sort(key=lambda t: t["start"])

        title = Text(title_text, font_size=26, weight=BOLD)
        subtitle = Text(subtitle_text, font_size=16, color=GRAY_B)
        header = VGroup(title, subtitle).arrange(DOWN, buff=0.2).to_corner(UL, buff=0.4)

        outer = Arc(
            radius=3.2,
            start_angle=PI * 0.2,
            angle=PI * 1.6,
            color=GRAY_B,
            stroke_width=2,
        )
        inner = Arc(
            radius=2.6,
            start_angle=PI * 0.2,
            angle=PI * 1.6,
            color=GRAY_C,
            stroke_width=1,
        )

        ticks = VGroup()
        for i in range(70):
            ang = outer.start_angle + (outer.angle * i / 69)
            p1 = 3.25 * np.array([np.cos(ang), np.sin(ang), 0])
            p2 = 3.4 * np.array([np.cos(ang), np.sin(ang), 0])
            ticks.add(Line(p1, p2, stroke_width=1, color=GRAY_B))

        points = VGroup()
        labels = VGroup()
        if dated:
            start_min = min(t["start"] for t in dated)
            end_max = max(t["start"] for t in dated)
        else:
            start_min = datetime.now()
            end_max = datetime.now()

        def date_to_angle(value: datetime) -> float:
            total = (end_max - start_min).days or 1
            offset = (value - start_min).days
            ratio = offset / total
            return outer.start_angle + outer.angle * ratio

        for idx, task in enumerate(dated):
            ang = date_to_angle(task["start"])
            radius = 2.9
            pos = radius * np.array([np.cos(ang), np.sin(ang), 0])
            points.add(Dot(pos, radius=0.05, color=BLUE_D))

            label = Text(f"ID {task['id']}", font_size=12, color=GRAY_B)
            label.move_to((radius + 0.45) * np.array([np.cos(ang), np.sin(ang), 0]))
            label.rotate(ang + PI / 2)
            labels.add(label)

        group = VGroup(outer, inner, ticks, points, labels)
        group.move_to(DOWN * 0.2)
        group.rotate(PI / 5, axis=RIGHT)
        group.rotate(-PI / 10, axis=UP)

        self.set_camera_orientation(phi=70 * DEGREES, theta=-35 * DEGREES)

        self.play(Write(header), run_time=1)
        self.play(Create(outer), Create(inner), run_time=1.2)
        self.play(LaggedStartMap(Create, ticks, lag_ratio=0.02), run_time=1.5)
        self.play(LaggedStartMap(FadeIn, points, lag_ratio=0.05), run_time=1.0)
        self.play(LaggedStartMap(FadeIn, labels, lag_ratio=0.05), run_time=1.0)
        self.play(Rotate(group, angle=PI / 2, axis=UP), run_time=4, rate_func=linear)
        self.wait(1)


if __name__ == "__main__":
    raise SystemExit(run_filter_cli())
