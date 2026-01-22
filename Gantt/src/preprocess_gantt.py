#!/usr/bin/env python3
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%y")
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%y")
        except ValueError:
            continue
    return text


def format_percent(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        pct = value * 100 if 0 <= value <= 1 else value
        return f"{int(round(pct))}%"
    text = str(value).strip()
    if not text:
        return ""
    return text if "%" in text else f"{text}%"


def parse_id_list(text: str) -> set[int]:
    ids: set[int] = set()
    if not text:
        return ids
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            left, right = part.split("-", 1)
            if left.strip().isdigit() and right.strip().isdigit():
                start = int(left.strip())
                end = int(right.strip())
                ids.update(range(min(start, end), max(start, end) + 1))
        elif part.isdigit():
            ids.add(int(part))
    return ids


def load_tasks_from_xlsx(path: Path) -> list[list]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in header_row]
    header_map = {str(val).strip().lower(): idx for idx, val in enumerate(headers, start=1) if val}

    def col_index(*names: str) -> int | None:
        for name in names:
            key = name.strip().lower()
            if key in header_map:
                return header_map[key]
        return None

    col_id = col_index("id", "uid", "uid ")
    col_name = col_index("nombre de la tarea", "nombre", "task name", "name")
    col_status = col_index("estado", "status")
    col_assigned = col_index("asignado a", "asignado", "assigned")
    col_start = col_index("fecha de inicio", "inicio", "start", "start date")
    col_end = col_index("fecha de finalización", "fecha de finalizacion", "fin", "finish", "end", "end date")
    col_pct = col_index("porcentaje completo", "% completo", "avance", "percent complete", "percentcomplete")
    col_duration = col_index("duración", "duracion", "duration")
    col_pred = col_index("predecesores", "predecessors", "pred")

    if not col_name:
        raise ValueError("No se encontro la columna 'Nombre de la tarea' en el XLSX.")

    tasks = []
    for row_idx in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=col_name)
        name_val = name_cell.value
        if name_val is None or str(name_val).strip() == "":
            continue

        indent = name_cell.alignment.indent or 0
        level = int(indent)

        task_id = ws.cell(row=row_idx, column=col_id).value if col_id else row_idx - 1
        status = ws.cell(row=row_idx, column=col_status).value if col_status else ""
        assigned = ws.cell(row=row_idx, column=col_assigned).value if col_assigned else ""
        start = ws.cell(row=row_idx, column=col_start).value if col_start else ""
        end = ws.cell(row=row_idx, column=col_end).value if col_end else ""
        pct = ws.cell(row=row_idx, column=col_pct).value if col_pct else ""
        duration = ws.cell(row=row_idx, column=col_duration).value if col_duration else ""
        pred = ws.cell(row=row_idx, column=col_pred).value if col_pred else ""

        if isinstance(task_id, float) and task_id.is_integer():
            task_id = int(task_id)

        tasks.append(
            [
                task_id,
                level,
                str(name_val).strip(),
                str(status).strip() if status is not None else "",
                str(assigned).strip() if assigned is not None else "",
                format_date(start),
                format_date(end),
                format_percent(pct),
                str(duration).strip() if duration is not None else "",
                str(pred).strip() if pred is not None else "",
            ]
        )

    return tasks


def build_parent_map(tasks: list[list]) -> dict[int, int | None]:
    parent_map: dict[int, int | None] = {}
    stack: list[tuple[int, int]] = []
    for row in tasks:
        task_id = row[0]
        level = row[1]
        while stack and stack[-1][1] >= level:
            stack.pop()
        parent_id = stack[-1][0] if stack else None
        if isinstance(task_id, int):
            parent_map[task_id] = parent_id
        stack.append((task_id, level))
    return parent_map


def build_children_map(parent_map: dict[int, int | None]) -> dict[int, list[int]]:
    children: dict[int, list[int]] = {}
    for child_id, parent_id in parent_map.items():
        if parent_id is None:
            continue
        children.setdefault(parent_id, []).append(child_id)
    return children


def collect_descendants(root_id: int, children: dict[int, list[int]]) -> set[int]:
    result = set()
    stack = [root_id]
    while stack:
        current = stack.pop()
        for child in children.get(current, []):
            if child not in result:
                result.add(child)
                stack.append(child)
    return result


def collect_ancestors(task_id: int, parent_map: dict[int, int | None]) -> list[int]:
    ancestors = []
    current = task_id
    while current in parent_map and parent_map[current] is not None:
        current = parent_map[current]
        ancestors.append(current)
    return ancestors


def filter_tasks(
    tasks: list[list],
    level: int | None,
    ids: set[int],
) -> tuple[list[list], list[str]]:
    warnings: list[str] = []
    id_to_row = {row[0]: row for row in tasks if isinstance(row[0], int)}
    parent_map = build_parent_map(tasks)
    children_map = build_children_map(parent_map)

    if level is not None:
        target_level = level + 1
    else:
        target_level = None

    if ids:
        missing = [i for i in ids if i not in id_to_row]
        if missing:
            warnings.append(f"IDs no encontrados: {', '.join(map(str, missing))}")

    selected_ids: set[int] = set()

    if ids:
        for task_id in sorted(ids):
            if task_id not in id_to_row:
                continue
            row = id_to_row[task_id]
            row_level = row[1]

            if level is not None and row_level != level:
                warnings.append(f"ID {task_id} no pertenece al nivel {level}.")
                continue

            has_children = task_id in children_map
            if has_children:
                selected_ids.add(task_id)
                selected_ids.update(collect_descendants(task_id, children_map))
            else:
                ancestors = collect_ancestors(task_id, parent_map)
                selected_ids.add(task_id)
                selected_ids.update(ancestors)

                parent_id = parent_map.get(task_id)
                if parent_id is not None:
                    siblings = children_map.get(parent_id, [])
                    selected_ids.update(siblings)
    elif target_level is not None:
        selected_ids = {row[0] for row in tasks if isinstance(row[0], int) and row[1] == target_level}
        if not selected_ids:
            warnings.append(f"No hay tareas en el nivel {target_level}.")
    else:
        selected_ids = {row[0] for row in tasks if isinstance(row[0], int)}

    filtered = [row for row in tasks if row[0] in selected_ids]
    return filtered, warnings


def print_stats(tasks: list[list], warnings: list[str]) -> None:
    levels = {}
    dated = 0
    for row in tasks:
        level = row[1]
        levels[level] = levels.get(level, 0) + 1
        if row[5] and row[6]:
            dated += 1

    if warnings:
        for msg in warnings:
            print(f"WARNING: {msg}", file=sys.stderr)

    print(f"Total tareas: {len(tasks)}", file=sys.stderr)
    print(f"Con fechas: {dated}", file=sys.stderr)
    for level in sorted(levels):
        print(f"Nivel {level}: {levels[level]}", file=sys.stderr)


def main() -> int:
    parser = argparse.ArgumentParser(description="Preprocesa tareas Gantt (solo XLSX).")
    parser.add_argument("-xlsx", "--xlsx", required=True, help="Ruta al archivo XLSX.")
    parser.add_argument("--nivel", type=int, default=None, help="Nivel a filtrar (ej: 2).")
    parser.add_argument("--id", dest="ids", default="", help="IDs separados por coma o rango (ej: 21,22,30-35).")
    args = parser.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(f"Error: no existe el archivo {path}", file=sys.stderr)
        return 1

    tasks = load_tasks_from_xlsx(path)
    ids = parse_id_list(args.ids)
    filtered, warnings = filter_tasks(tasks, args.nivel, ids)
    print_stats(filtered, warnings)

    print("tasks = [")
    for row in filtered:
        print(f"    {row},")
    print("]")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
